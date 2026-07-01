import streamlit as st
import openpyxl
import csv
import io
from datetime import date, timedelta
from pathlib import Path

st.set_page_config(page_title="Secondary Market Summary", page_icon="📊", layout="centered")

# ── Styling ───────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.section-title { font-size: 1.1rem; font-weight: 600; color: #0f172a; margin: 2rem 0 0.25rem; }
.section-sub   { font-size: 0.82rem; color: #64748b; margin-bottom: 1rem; }
.summary-table { width: 100%; border-collapse: collapse; margin: 1rem 0; }
.summary-table th {
    background: #0f172a; color: #fff;
    padding: 9px 14px; text-align: right;
    font-size: 0.82rem; font-weight: 500; letter-spacing: 0.04em;
}
.summary-table th:first-child { text-align: left; }
.summary-table td {
    padding: 9px 14px; text-align: right;
    border-bottom: 1px solid #e2e8f0; font-size: 0.88rem; color: #1e293b;
}
.summary-table td:first-child { text-align: left; color: #475569; font-weight: 500; }
.summary-table tr:last-child td { border-bottom: none; font-weight: 600; }
.dash { color: #cbd5e1 !important; }
.week-badge {
    display: inline-block; background: #f0fdf4; border: 1px solid #bbf7d0;
    border-radius: 6px; padding: 4px 12px; font-size: 0.8rem; color: #166534;
    margin-bottom: 1rem;
}
.file-badge {
    display: inline-block; background: #f1f5f9; border: 1px solid #e2e8f0;
    border-radius: 6px; padding: 4px 10px; font-size: 0.8rem; color: #475569; margin: 3px;
}
hr.divider { border: none; border-top: 1px solid #e2e8f0; margin: 2rem 0; }
</style>
""", unsafe_allow_html=True)

AYL_CELL = '\u0531\u0545\u053c'
CURRENCIES = ['AMD', 'USD', 'EUR', 'RUB', 'AYL']


# ── Helpers ───────────────────────────────────────────────────

def get_previous_week_range(today=None):
    if today is None:
        today = date.today()
    this_monday = today - timedelta(days=today.weekday())
    prev_monday = this_monday - timedelta(weeks=1)
    prev_friday = prev_monday + timedelta(days=4)
    return prev_monday, prev_friday


def parse_num(s):
    if s is None:
        return None
    try:
        return float(str(s).strip())
    except:
        return None


def parse_date_dmy(s):
    from datetime import datetime
    try:
        return datetime.strptime(str(s).strip(), '%d/%m/%Y').date()
    except:
        return None


def weighted_avg(values, weights):
    total_w = sum(weights)
    if total_w == 0:
        return None
    return sum(v * w for v, w in zip(values, weights)) / total_w


def fmt_num(n):
    return '\u2014' if not n else f"{n:,.0f}"

def fmt_pct(p):
    return '\u2014' if p is None else f"{p:.4f}%"

def fmt_rate(p):
    return '\u2014' if p is None else f"{p:.4f}"


def summary_table_html(rows, headers):
    th = '<tr>' + ''.join(f'<th>{h}</th>' for h in headers) + '</tr>'
    body = ''
    for row in rows:
        body += '<tr>' + ''.join(
            f'<td class="dash">{v}</td>' if v == '\u2014' else f'<td>{v}</td>'
            for v in row
        ) + '</tr>'
    return f'<table class="summary-table"><thead>{th}</thead><tbody>{body}</tbody></table>'


# ── Section 1: 6.2-Repo (xlsx) ───────────────────────────────

def get_merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for mc in ws.merged_cells.ranges:
        if mc.min_col <= col <= mc.max_col and mc.min_row <= row <= mc.max_row:
            return ws.cell(row=mc.min_row, column=mc.min_col).value
    return None


def parse_repo_sheet(ws):
    COL_C, COL_F, COL_H, COL_L, COL_N = 3, 6, 8, 12, 14
    label_map = {'AMD': 'AMD', 'USD': 'USD', 'EUR': 'EUR', 'RUB': 'RUB', AYL_CELL: 'AYL'}
    data = {cur: {'F': [], 'H': [], 'L': [], 'N': []} for cur in CURRENCIES}
    for row_num in range(1, ws.max_row + 1):
        raw = get_merged_value(ws, row_num, COL_C)
        currency = label_map.get(raw)
        if currency is None:
            continue
        f = ws.cell(row=row_num, column=COL_F).value
        h = ws.cell(row=row_num, column=COL_H).value
        l = ws.cell(row=row_num, column=COL_L).value
        n = ws.cell(row=row_num, column=COL_N).value
        if isinstance(f, (int, float)) and f > 0 and isinstance(h, (int, float)):
            data[currency]['F'].append(f)
            data[currency]['H'].append(h)
        if isinstance(l, (int, float)) and l > 0 and isinstance(n, (int, float)):
            data[currency]['L'].append(l)
            data[currency]['N'].append(n)
    return data


def process_repo_xlsx(uploaded_files):
    accum = {cur: {'F_vals': [], 'H_vals': [], 'L_vals': [], 'N_vals': []} for cur in CURRENCIES}
    errors = []
    for uf in uploaded_files:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uf.read()), data_only=True)
        except Exception as e:
            errors.append(f"{uf.name}: {e}")
            continue
        if '6.2-Repo' not in wb.sheetnames:
            errors.append(f"{uf.name}: Sheet '6.2-Repo' not found.")
            continue
        ws = wb['6.2-Repo']
        day_data = parse_repo_sheet(ws)
        for cur in CURRENCIES:
            accum[cur]['F_vals'].extend(day_data[cur]['F'])
            accum[cur]['H_vals'].extend(day_data[cur]['H'])
            accum[cur]['L_vals'].extend(day_data[cur]['L'])
            accum[cur]['N_vals'].extend(day_data[cur]['N'])
    results = {}
    for cur in CURRENCIES:
        fl = accum[cur]['F_vals']; hl = accum[cur]['H_vals']
        ll = accum[cur]['L_vals']; nl = accum[cur]['N_vals']
        results[cur] = {
            'total': sum(fl) + sum(ll),
            'total_F': sum(fl), 'total_L': sum(ll),
            'wavg': weighted_avg(hl + nl, fl + ll),
        }
    return results, errors


# ── Section 2: Repo/Reverse Repo CSV ─────────────────────────

def process_auction_csv(file_bytes, prev_monday, prev_friday):
    """
    File 1 columns (0-indexed):
      [5]  = date (Կнqman amsаtiw)
      [16] = Declared Volume (Ughenishayiн tsaval)   ← col Q
      [4]  = Satisfied Volume                         ← col E
      [13] = Weighted avg rate                        ← col N
    """
    text = file_bytes.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    next(reader)  # skip header

    declared = []    # col Q [16]
    satisfied = []   # col E [4]
    rates_q = []     # rates paired with declared
    rates_e = []     # rates paired with satisfied

    for row in reader:
        if len(row) < 17:
            continue
        d = parse_date_dmy(row[5])
        if d is None or not (prev_monday <= d <= prev_friday):
            continue
        q = parse_num(row[16])
        e = parse_num(row[4])
        n = parse_num(row[13])
        if q and q > 0 and n is not None:
            declared.append(q)
            rates_q.append(n)
        if e and e > 0 and n is not None:
            satisfied.append(e)
            rates_e.append(n)

    return {
        'declared':  sum(declared),
        'satisfied': sum(satisfied),
        'wavg_q':    weighted_avg(rates_q, declared),
        'wavg_e':    weighted_avg(rates_e, satisfied),
    }


# ── Section 3: Foreign Exchange CSV ──────────────────────────

def process_fx_csv(file_bytes, prev_monday, prev_friday):
    """
    File 2 columns (0-indexed):
      [5]  = date
      [8]  = Declared Volume (Ughenishayiн tsaval)   ← col I
      [4]  = Satisfied Volume                         ← col E
      [11] = Weighted avg exchange rate               ← col L
    """
    text = file_bytes.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    next(reader)  # skip header

    declared = []
    satisfied = []
    rates_i = []
    rates_e = []

    for row in reader:
        if len(row) < 12:
            continue
        d = parse_date_dmy(row[5])
        if d is None or not (prev_monday <= d <= prev_friday):
            continue
        i_val = parse_num(row[8])
        e_val = parse_num(row[4])
        l_val = parse_num(row[11])
        if i_val and i_val > 0 and l_val is not None:
            declared.append(i_val)
            rates_i.append(l_val)
        if e_val and e_val > 0 and l_val is not None:
            satisfied.append(e_val)
            rates_e.append(l_val)

    return {
        'declared':  sum(declared),
        'satisfied': sum(satisfied),
        'wavg_i':    weighted_avg(rates_i, declared),
        'wavg_e':    weighted_avg(rates_e, satisfied),
    }


# ═══════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════

st.markdown('<div class="section-title" style="font-size:1.4rem;margin-top:0">📊 Secondary Market Summary</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub" style="margin-bottom:1.5rem">Upload your files below to generate summaries.</div>', unsafe_allow_html=True)

prev_monday, prev_friday = get_previous_week_range()
week_label = f"Previous week: {prev_monday.strftime('%b %d')} – {prev_friday.strftime('%b %d, %Y')}"
st.markdown(f'<div class="week-badge">📅 {week_label}</div>', unsafe_allow_html=True)


# ── Block 1 ───────────────────────────────────────────────────
st.markdown('<div class="section-title">1 · Repo Operations (6.2-Repo sheet)</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Upload one .xlsx file per day (up to 5 for the full week).</div>', unsafe_allow_html=True)

repo_files = st.file_uploader("Repo xlsx files", type=["xlsx"],
    accept_multiple_files=True, label_visibility="collapsed", key="repo")

if repo_files:
    badges = "".join(f'<span class="file-badge">📄 {f.name}</span>' for f in repo_files)
    st.markdown(badges, unsafe_allow_html=True)
    results, errors = process_repo_xlsx(repo_files)
    for e in errors:
        st.error(e)
    cols = ['AMD', 'EUR', 'USD', 'RUB']
    rows = [
        ['Banks and Financial Institutions'] + [fmt_num(results[c]['total']) for c in cols],
        ['Weighted Avg Rate']               + [fmt_rate(results[c]['wavg'])  for c in cols],
    ]
    st.markdown(summary_table_html(rows, [''] + cols), unsafe_allow_html=True)
    ayl = results.get('AYL', {})
    if ayl.get('total', 0) > 0:
        st.info(f"AYL: {fmt_num(ayl['total'])}  |  Weighted Avg: {fmt_rate(ayl['wavg'])}")
else:
    st.info("No files uploaded yet.")


st.markdown('<hr class="divider">', unsafe_allow_html=True)


# ── Block 2 ───────────────────────────────────────────────────
st.markdown('<div class="section-title">2 · Repo / Reverse Repo / Auction</div>', unsafe_allow_html=True)
st.markdown(f'<div class="section-sub">Upload the CSV. Rows are automatically filtered to {week_label.lower()}.</div>', unsafe_allow_html=True)

auction_file = st.file_uploader("Auction CSV", type=["csv"],
    accept_multiple_files=False, label_visibility="collapsed", key="auction")

if auction_file:
    st.markdown(f'<span class="file-badge">📄 {auction_file.name}</span>', unsafe_allow_html=True)
    res = process_auction_csv(auction_file.read(), prev_monday, prev_friday)
    if res['declared'] == 0 and res['satisfied'] == 0:
        st.warning(f"No data found for {week_label.lower()}. Check the file covers that date range.")
    else:
        rows = [
            ['Declared Volume (col Q)',  fmt_num(res['declared']),  fmt_rate(res['wavg_q'])],
            ['Satisfied Volume (col E)', fmt_num(res['satisfied']), fmt_rate(res['wavg_e'])],
        ]
        st.markdown(summary_table_html(rows, ['', 'Total', 'Weighted Avg Rate']), unsafe_allow_html=True)
else:
    st.info("No file uploaded yet.")


st.markdown('<hr class="divider">', unsafe_allow_html=True)


# ── Block 3 ───────────────────────────────────────────────────
st.markdown('<div class="section-title">3 · Foreign Exchange (Buy / Sell)</div>', unsafe_allow_html=True)
st.markdown(f'<div class="section-sub">Upload the CSV. Rows are automatically filtered to {week_label.lower()}.</div>', unsafe_allow_html=True)

fx_file = st.file_uploader("FX CSV", type=["csv"],
    accept_multiple_files=False, label_visibility="collapsed", key="fx")

if fx_file:
    st.markdown(f'<span class="file-badge">📄 {fx_file.name}</span>', unsafe_allow_html=True)
    res = process_fx_csv(fx_file.read(), prev_monday, prev_friday)
    if res['declared'] == 0 and res['satisfied'] == 0:
        st.warning(f"No data found for {week_label.lower()}. Check the file covers that date range.")
    else:
        rows = [
            ['Declared Volume (col I)',   fmt_num(res['declared']),  fmt_rate(res['wavg_i'])],
            ['Satisfied Volume (col E)',  fmt_num(res['satisfied']), fmt_rate(res['wavg_e'])],
        ]
        st.markdown(summary_table_html(rows, ['', 'Total', 'Weighted Avg Rate']), unsafe_allow_html=True)
else:
    st.info("No file uploaded yet.")
